import vk_api
from vk_api.longpoll import VkLongPoll, VkEventType
from vk_api.utils import get_random_id
import re
import csv
import io
import os
import json
import psycopg2
from psycopg2.extras import RealDictCursor

# --- ФУНКЦИЯ ОБРЕЗКИ ТЕКСТА ---
def truncate_label(text, max_len=40):
    if len(text) <= max_len:
        return text
    return text[:max_len - 3] + "…"
    
def format_numbered_list(items, start_from=1):
    """Делает нумерованный список: '1 — вариант', '2 — вариант' и т.д."""
    lines = []
    for i, item in enumerate(items, start=start_from):
        # Если строка очень длинная — обрежем, чтобы не ломать читаемость
        if len(item) > 80:
            item = item[:77] + "…"
        lines.append(f"{i} — {item}")
    return "\n".join(lines)

def parse_numbers_input(text, items, start_from=1):
    """
    Превращает ввод пользователя '1', '1, 3', '2' в реальные значения из списка.
    Возвращает список значений или None, если ввод некорректен.
    """
    parts = [p.strip() for p in text.split(",") if p.strip()]
    if not parts:
        return None

    selected = []
    for p in parts:
        if not p.isdigit():
            return None  # не цифра
        idx = int(p) - start_from
        if idx < 0 or idx >= len(items):
            return None  # номер вне диапазона
        selected.append(items[idx])

    return selected
# -----------------------------------------------------

# ================= НАСТРОЙКИ ИЗ ENV =================
VK_TOKEN = os.getenv("VK_TOKEN")
ADMIN_IDS = set(map(int, os.getenv("ADMIN_IDS", "").split(","))) if os.getenv("ADMIN_IDS") else set()
GROUP_ID = int(os.getenv("GROUP_ID"))
DATABASE_URL = os.getenv("DATABASE_URL")

if not VK_TOKEN or not DATABASE_URL:
    raise ValueError("Не заданы переменные окружения VK_TOKEN или DATABASE_URL")
# =====================================================
vk_session = vk_api.VkApi(token=VK_TOKEN)
vk = vk_session.get_api()
longpoll = VkLongPoll(vk_session)
# ----------------- БАЗА ДАННЫХ -----------------
def get_db():
    return psycopg2.connect(DATABASE_URL)

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS answers (
            user_id BIGINT PRIMARY KEY,
            fio TEXT, institution TEXT, specialty TEXT, study_group TEXT,
            course TEXT, form_of_study TEXT, contacts TEXT,
            employment_status TEXT, target_contract TEXT, experience TEXT,
            practice_eval TEXT, events TEXT, resume_status TEXT,
            interview_training TEXT, special_status TEXT, military TEXT,
            maternity TEXT, graduate TEXT, post_plans TEXT, help_needed TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS progress (
            user_id BIGINT PRIMARY KEY,
            step_index INTEGER DEFAULT 0,
            uni_page INTEGER DEFAULT 0,
            started INTEGER DEFAULT 0
        )
    ''')
    conn.commit()
    conn.close()

def get_progress(user_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT step_index, uni_page, started FROM progress WHERE user_id=%s", (user_id,))
    row = c.fetchone()
    conn.close()
    return (row[0], row[1], row[2]) if row else (0, 0, 0)

def set_progress(user_id, step_index, uni_page=0, started=1):
    conn = get_db()
    c = conn.cursor()
    c.execute(
        "INSERT INTO progress (user_id, step_index, uni_page, started) VALUES (%s,%s,%s,%s) "
        "ON CONFLICT (user_id) DO UPDATE SET step_index=%s, uni_page=%s, started=%s",
        (user_id, step_index, uni_page, started, step_index, uni_page, started)
    )
    conn.commit()
    conn.close()

def save_answer(user_id, field, value):
    cols = ["fio","institution","specialty","study_group","course","form_of_study",
            "contacts","employment_status","target_contract","experience",
            "practice_eval","events","resume_status","interview_training",
            "special_status","military","maternity","graduate","post_plans","help_needed"]
    if field not in cols:
        return
    conn = get_db()
    c = conn.cursor()
    c.execute("INSERT INTO answers (user_id) VALUES (%s) ON CONFLICT (user_id) DO NOTHING", (user_id,))
    c.execute(f'UPDATE answers SET {field}=%s WHERE user_id=%s', (value, user_id))
    conn.commit()
    conn.close()

# ----------------- СПИСОК ВУЗов -----------------
UNIVERSITIES = [
    "БПОУ УР «Воткинский промышленный техникум»", "БПОУ УР «Воткинский музыкально-педагогический колледж имени П.И. Чайковского»",
    "БПОУ УР «Воткинский машиностроительный техникум имени В.Г.Садовникова»", "АПОУ УР «Глазовский аграрно-промышленный техникум»",
    "БПОУ УР «Глазовский технический колледж»", "БПОУ «Дебесский политехникум»", "БПОУР «Игринский политехнический техникум»",
    "БПОУ УР «Ижевский торгово-экономический техникум»", "БПОУ УР «Ижевский монтажный техникум»", "БПОУ «Ижевский агростроительный техникум»",
    "ЧПОО «Нефтяной техникум»", "БПОУ УР «Ижевский политехнический колледж»", "БПОУ УР «Ижевский промышленно-экономический колледж»",
    "БПОУ УР «Ижевский машиностроительный техникум им. С.Н. Борина»", "БПОУ УР «Радиомеханический техникум имени В.А. Шутова»",
    "АПОУ УР «Экономико-технологический колледж»", "БПОУ УР «Асановский аграрно-технический техникум»", "АПОУ УР «Топливно-энергетический колледж»",
    "БПОУ УР «Ижевский техникум индустрии питания»", "КПОУ УР «Удмуртский республиканский колледж культуры»",
    "АНПОО «Международный Восточно-Европейский колледж»", "АПОУ УР «Техникум радиоэлектроники и информационных технологий им. А.В. Воскресенского»",
    "АПОУ УР «Республиканский медицинский колледж имени Героя Советского Союза Ф.А. Пушиной Министерства здравоохранения Удмуртской Республики»",
    "ПОЧУ «Ижевский техникум экономики, управления и права Удмуртпотребсоюза»", "АНПОО СПО «Ижевский финансово-юридический колледж»",
    "БПОУ УР «Удмуртский республиканский социально-педагогический колледж»", "АПОУ УР «Строительный техникум»",
    "ФГБОУ ВО «Ижевская государственная медицинская академия»", "КПОУ УР «Республиканский музыкальный колледж»",
    "БПОУ УР «Ижевский промышленно-экономический колледж» в г. Можга", "БПОУ УР «Можгинский педагогический колледж имени Т.К. Борисова»",
    "БПОУ УР «Можгинский агропромышленный колледж»", "БПОУ УР «Сарапульский политехнический техникум»",
    "БПОУ УР «Сарапульский многопрофильный колледж»", "БПОУ УР «Сарапульский колледж социально-педагогических технологий и сервиса»",
    "БПОУ УР «Сюмсинский техникум лесного и сельского хозяйства»", "БПОУ УР «Увинский профессиональный колледж»",
    "БПОУ УР «Ярский политехникум»", "ФГБОУ ВО «Приволжский государственный университет путей сообщения»",
    "БПОУ УР «Ижевский индустриальный техникум имени Евгения Фёдоровича Драгунова»", "БПОУ УР «Глазовский политехнический колледж»",
    "Министерство юстиции Российской Федерации", "ФГБОУ ВО «Удмуртский государственный университет»", "ФГБОУ ВО «Удмуртский государственный аграрный университет»", 
    "ФГБОУ ВО «Ижевский государственный технический университет имени М.Т. Калашникова»", "БПОУ УР «Ижевский автотранспортный техникум»", 
    "ФГБОУ ВО «Глазовский государственный инженерно-педагогический университет имени В. Г. Короленко»", "Сарапульский техникум машиностроения и информационных технологий"
]
ITEMS_PER_PAGE = 10

# ----------------- ШАГИ АНКЕТЫ -----------------
STEPS = [
    "fio", "institution", "specialty", "study_group",
    "course", "form_of_study", "contacts",
    "employment_status", "target_contract", "experience",
    "practice_eval", "events", "resume_status",
    "interview_training", "special_status", "military",
    "maternity", "graduate", "post_plans", "help_needed"
]

STEP_TO_DB = {
    "fio": "fio", "institution": "institution", "specialty": "specialty",
    "study_group": "study_group", "course": "course", "form_of_study": "form_of_study",
    "contacts": "contacts", "employment_status": "employment_status",
    "target_contract": "target_contract", "experience": "experience",
    "practice_eval": "practice_eval", "events": "events",
    "resume_status": "resume_status", "interview_training": "interview_training",
    "special_status": "special_status", "military": "military",
    "maternity": "maternity", "graduate": "graduate",
    "post_plans": "post_plans", "help_needed": "help_needed"
}

QUESTIONS = {
    "fio": "Пожалуйста, укажите ваши фамилию, имя и отчество полностью:",
    "institution": "Выберите ваше учебное заведение из списка (используйте «Далее →» для пролистывания):",
    "specialty": "Укажите вашу специальность обучения:",
    "study_group": "Укажите номер вашей учебной группы:",
    "course": "Выберите ваш курс обучения:",
    "form_of_study": "Выберите форму обучения:",
    "contacts": "Укажите контактные данные — телефон или e-mail (например: +79991234567 или student@mail.ru):",
    "employment_status": (
        "Ваш статус занятости прямо сейчас.\n"
        "Выберите один или несколько вариантов, указав их номера через запятую (например: 1, 3):\n\n"
        "1 — Работаю по трудовому договору (в том числе по совместительству)\n"
        "2 — Работаю по гражданско-правовому договору (договор подряда, услуг и т.п.)\n"
        "3 — Являюсь самозанятым / ИП / учредителем юрлица\n"
        "4 — Прохожу оплачиваемую стажировку / практику у работодателя\n"
        "5 — Работаю временно (разовые подработки), не по специальности обучения\n"
        "6 — Ничего из вышеперечисленного"
    ),
    "target_contract": "Есть ли у вас заключённый договор о целевом обучении с работодателем?",
    "experience": "Есть ли у вас опыт работы или оплачиваемой стажировки по основной или близкой к ней специальности?",
    "practice_eval": "Как вы в целом оцениваете результаты своих производственных практик у работодателей по специальности обучения?",
    "events": (
        "Участвовали ли вы в течение обучения в мероприятиях, которые помогают познакомиться с работодателями "
        "(ярмарки вакансий, дни карьеры, профтуры на предприятия, встречи с работодателями и т.д.)?"
    ),
    "resume_status": "Наличие резюме для поиска работы:",
    "interview_training": "Проходили ли вы занятия или тренинги по навыкам прохождения собеседования?",
    "special_status": "Есть ли у вас особый статус или жизненные обстоятельства?",
    "military": (
        "Планируется ли в отношении вас призыв на военную службу в ближайшее время "
        "(после окончания текущего года обучения)?"
    ),
    "maternity": (
        "Планируете ли вы уходить в отпуск по уходу за ребёнком "
        "в период обучения или сразу после окончания обучения (или продолжать уже начатый отпуск)?"
    ),
    "graduate": "Являетесь ли вы студентом выпускного курса (оканчиваете программу в текущем учебном году)?",
    "post_plans": (
        "Ваши планы после выпуска.\n"
        "Выберите один или несколько вариантов, указав их номера через запятую (например: 1, 3):\n\n"
        "1 — У меня есть подписанный трудовой договор (или договор на целевое обучение)\n"
        "2 — Есть устная договорённость с работодателем, но без подписанных документов\n"
        "3 — Прохожу стажировку\n"
        "4 — Планирую организовать своё дело (самозанятость / ИП / учредитель юрлица)\n"
        "5 — Планирую продолжить обучение (магистратура / аспирантура и пр.)\n"
        "6 — Сейчас ищу работу\n"
        "7 — Пока нет планов"
    ),
    "help_needed": (
        "Какую помощь от Кадрового центра «Работа России» вы бы считали наиболее полезной?\n"
        "Выберите все подходящие варианты, указав их номера через запятую (например: 1, 2, 4):\n\n"
        "1 — Подбор актуальных вакансий с учётом специальности\n"
        "2 — Тренинги по составлению резюме, подготовке к собеседованиям, сопроводительных писем\n"
        "3 — Профтур-экскурсии на предприятия\n"
        "4 — Подбор оплачиваемой стажировки\n"
        "5 — Подбор работодателя для практики\n"
        "6 — Заключение договора с работодателем на целевое обучение\n"
        "7 — Помощь с ЕЦП «Работа России»\n"
        "8 — Другое (укажите)"
    )
}

OPTIONS = {
    "course": ["1 курс", "2 курс", "3 курс", "4 курс", "5 курс"],
    "form_of_study": ["очная", "очно-заочная", "заочная"],
    "target_contract": [
        "да, договор о целевом обучении заключён",
        "нет, договора о целевом обучении нет"
    ],
    "experience": [
        "да, есть опыт работы / оплачиваемой стажировки по основной или близкой специальности",
        "есть опыт работы только вне специальности обучения",
        "нет, опыта работы и оплачиваемых стажировок пока не было"
    ],
    "practice_eval": [
        "скорее доволен(льна) или полностью доволен(льна)",
        "скорее не доволен(льна) / совсем не доволен(льна) результатами практик"
    ],
    "events": [
        "да, за последний год участвовал(а) хотя бы в одном таком мероприятии",
        "участвовал(а), но более года назад",
        "нет, ещё ни разу не участвовал(а)"
    ],
    "resume_status": [
        "есть актуальное резюме, которым я пользуюсь или готов(а) пользоваться",
        "резюме есть, но оно устарело / резюме нет, я его не составлял(а)"
    ],
    "interview_training": [
        "да, проходил(а) одно или несколько таких мероприятий",
        "пока не проходил(а)"
    ],
    "special_status": [
        "да, имею группу инвалидности",
        "отношусь к категории детей-сирот и детей, оставшихся без попечения родителей",
        "планирую переезд в другой регион / страну после окончания обучения",
        "ничего из вышеперечисленного"
    ],
    "military": [
        "да, планируется призыв",
        "нет / не подлежу призыву / вопрос уже решён (служба пройдена и др.)"
    ],
    "maternity": ["да, планирую", "пока не планирую"],
    "graduate": [
        "да, я учусь на выпускном курсе",
        "нет, я не на выпускном курсе"
    ]
}

MULTI_STEPS = ["employment_status", "post_plans", "help_needed"]

MESSAGES = {
    "welcome": (
        "👋 Здравствуйте! Я бот для оценки рисков нетрудоустройства студентов.\n\n"
        "Я задам несколько коротких вопросов о вашем обучении и занятости. "
        "Это займёт примерно 5–7 минут. Ответы помогут подобрать для вас "
        "подходящие вакансии, стажировки и поддержку от центра занятости.\n\n"
        "Нажмите кнопку «Начать анкету», чтобы приступить."
    ),
    "invalid_contact": (
        "Не удалось распознать контакт. Пожалуйста, введите:\n\n"
        "• Номер телефона в формате +79991234567 или 89991234567\n"
        "или\n"
        "• Адрес электронной почты в формате example@mail.ru"
    ),
    "invalid_choice": "Пожалуйста, выберите один из вариантов, нажав на кнопку ниже.",
    "invalid_multi": "Пожалуйста, укажите номера вариантов через запятую (например: 1, 3, 5). Проверьте, что номера от 1 до {}.",
    "no_data": "Пока нет собранных анкет для выгрузки.",
    "admin_only": "Эта команда доступна только администраторам.",
    "finished": (
        "✅ Спасибо! Анкета заполнена.\n\n"
        "Если у вас возникнут вопросы, вы всегда можете написать сюда ещё раз."
    )
}

# ----------------- ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ -----------------

def export_to_table(admin_id):
    conn = get_db()
    c = conn.cursor(cursor_factory=RealDictCursor)
    c.execute("SELECT * FROM answers")
    rows = c.fetchall()
    conn.close()

    if not rows:
        send_message(admin_id, MESSAGES["no_data"])
        return

    # ---------- Генерируем .xlsx через openpyxl ----------
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Анкеты"

    cols = list(rows[0].keys())

    # Заголовки жирным
    header_font = Font(bold=True)
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Данные
    for row_idx, r in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(cols, start=1):
            val = r[col_name]
            ws.cell(row=row_idx, column=col_idx, value=val if val is not None else "")

    # Авто-ширина колонок (примерная)
    for col_idx, col_name in enumerate(cols, start=1):
        max_len = max(len(str(col_name)), max(
            (len(str(r[col_name])) if r[col_name] else 0) for r in rows
        ))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)

    fname = "survey_export.xlsx"
    wb.save(fname)

    # ---------- Пытаемся загрузить .xlsx в ВК ----------
    try:
        # Получаем адрес сервера для загрузки
        upload_server = vk.docs.getMessagesUploadServer(
            type="doc",
            peer_id=admin_id
        )
        upload_url = upload_server["upload_url"]

        # Загружаем файл на сервер ВК
        import urllib.request
        import urllib.error
        import mimetypes
        from io import BytesIO

        boundary = "----WebKitFormBoundary7MA4YWxkTrZu0gW"
        with open(fname, "rb") as f:
            file_data = f.read()

        body = (
            f"--{boundary}\r\n"
            f'Content-Disposition: form-data; name="file"; filename="survey_export.xlsx"\r\n'
            f"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
        ).encode("utf-8") + file_data + f"\r\n--{boundary}--\r\n".encode("utf-8")

        req = urllib.request.Request(
            upload_url,
            data=body,
            headers={"Content-Type": f"multipart/form-data; boundary={boundary}"}
        )
        resp = urllib.request.urlopen(req)
        import json as _json
        result = _json.loads(resp.read().decode("utf-8"))

        if "file" not in result or not result["file"]:
            raise Exception("Сервер ВК не принял файл")

        # Сохраняем документ в ВК
        doc = vk.docs.save(file=result["file"], title="Выгрузка анкет.xlsx")

        # Достаём owner_id и id документа
        if isinstance(doc, dict) and "doc" in doc:
            d = doc["doc"]
        elif isinstance(doc, dict) and "docs" in doc:
            d = doc["docs"][0]
        else:
            raise Exception(f"Неожиданный ответ docs.save: {doc}")

        att = f"doc{d['owner_id']}_{d['id']}"
        send_message(admin_id, "📊 Вот выгрузка анкет в Excel:", attachment=att)

    except Exception as e:
        print(f"Ошибка загрузки .xlsx в ВК: {e}")

        # ---------- Fallback: отправляем CSV текстом ----------
        try:
            out = io.StringIO()
            writer = csv.DictWriter(out, fieldnames=cols)
            writer.writeheader()
            for r in rows:
                writer.writerow(dict(r))
            csv_text = out.getvalue()
            out.close()

            if len(csv_text) > 4000:
                # Разбиваем на части
                chunks = []
                lines = csv_text.split("\n")
                current = ""
                for line in lines:
                    if len(current) + len(line) + 1 > 4000:
                        chunks.append(current)
                        current = line + "\n"
                    else:
                        current += line + "\n"
                if current:
                    chunks.append(current)

                for i, chunk in enumerate(chunks):
                    header = f"📊 Выгрузка анкет (часть {i+1}/{len(chunks)}):\n\n"
                    send_message(admin_id, header + chunk)
            else:
                send_message(admin_id, "📊 Выгрузка анкет (CSV, откроется в Excel):\n\n" + csv_text)
        except Exception as e2:
            send_message(admin_id, f"Не удалось выгрузить данные: {e2}")

    finally:
        if os.path.exists(fname):
            os.remove(fname)

def send_message(user_id, message, keyboard=None, attachment=None):
    try:
        params = {
            "peer_id": user_id,        
            "message": message,
            "random_id": get_random_id()
        }
        if keyboard:
            params["keyboard"] = keyboard
        if attachment:
            params["attachment"] = attachment
        vk.messages.send(**params)
    except Exception as e:
        print(f"Ошибка отправки: {e}")

def ask_step_with_numbers(user_id, step_key):
    """Отправляет вопрос и нумерованный список вариантов (без кнопок)."""
    question = QUESTIONS[step_key]
    options = OPTIONS[step_key]

    list_text = format_numbered_list(options)
    message = f"{question}\n\n{list_text}\n\nНапишите номер(а) выбранного варианта(ов) через запятую (например: 1, 3)."

    send_message(user_id, message)

def ask_university_page(user_id, page):
    start = page * ITEMS_PER_PAGE
    end = min(start + ITEMS_PER_PAGE, len(UNIVERSITIES))
    items = UNIVERSITIES[start:end]

    if not items:
        send_message(user_id, "Список вузов пуст.")
        return

    list_text = format_numbered_list(items, start_from=start + 1)

    nav = []
    if page > 0:
        nav.append("← 0 — Назад к началу списка")
    if end < len(UNIVERSITIES):
        nav.append(f"→ {len(UNIVERSITIES) // ITEMS_PER_PAGE + 1} — Следующая страница")

    nav_text = "\n".join(nav) if nav else ""

    message = (
        "Выберите ваш вуз из списка:\n\n"
        f"{list_text}\n\n"
        f"{nav_text}\n\n"
        "Введите номер вуза (например: 5). Если нужен другой список — используйте номера навигации."
    )
    send_message(user_id, message)

def parse_numbered_input(text, items, start_from=1):
    """
    text: строка вида '1', '1, 3', '2' и т.п.
    items: список вариантов
    возвращает список выбранных значений или None, если ввод некорректен
    """
    parts = [p.strip() for p in text.split(",") if p.strip()]
    if not parts:
        return None

    selected = []
    for p in parts:
        if not p.isdigit():
            return None  # не цифра — ошибка
        idx = int(p) - start_from
        if idx < 0 or idx >= len(items):
            return None  # номер вне диапазона
        selected.append(items[idx])

    return selected

    nav = []
    if page > 0:
        nav.append({"action": {"type": "text", "label": "← Назад"}, "color": "secondary"})
    if end < len(UNIVERSITIES):
        nav.append({"action": {"type": "text", "label": "Далее →"}, "color": "secondary"})
    if nav:
        rows.append(nav)

    return json.dumps({"one_time": False, "buttons": rows})


def validate_contact(text):
    text = text.strip()
    phone_re = r'^\+?[78]?[\s\-]?\(?\d{3}\)?[\s\-]?\d{3}[\s\-]?\d{2}[\s\-]?\d{2}$'
    email_re = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
    if re.match(phone_re, text):
        digits = re.sub(r'\D', '', text)
        if len(digits) == 11 and digits.startswith('8'):
            digits = '7' + digits[1:]
        elif len(digits) == 10:
            digits = '7' + digits
        return True, '+7' + digits[-10:]
    if re.match(email_re, text):
        return True, text.lower()
    return False, None

def parse_multi_choice(text, max_opt):
    try:
        parts = [p.strip() for p in text.split(',') if p.strip()]
        nums = [int(p) for p in parts]
        if any(n < 1 or n > max_opt for n in nums):
            return None
        return nums
    except ValueError:
        return None

def multi_to_text(step_key, nums):
    labels = {
        "employment_status": [
            "работаю по трудовому договору (в т.ч. по совместительству)",
            "работаю по гражданско-правовому договору",
            "самозанятый / ИП / учредитель юрлица",
            "прохожу оплачиваемую стажировку / практику",
            "работаю временно, не по специальности",
            "ничего из вышеперечисленного"
        ],
        "post_plans": [
            "есть подписанный трудовой договор / целевое обучение",
            "есть устная договорённость с работодателем",
            "прохожу стажировку",
            "планирую организовать своё дело",
            "планирую продолжить обучение",
            "сейчас ищу работу",
            "пока нет планов"
        ],
        "help_needed": [
            "подбор актуальных вакансий",
            "тренинги по резюме и собеседованиям",
            "профтур-экскурсии на предприятия",
            "подбор оплачиваемой стажировки",
            "подбор работодателя для практики",
            "заключение договора на целевое обучение",
            "помощь с ЕЦП «Работа России»",
            "другое"
        ]
    }
    return "; ".join(labels[step_key][n-1] for n in nums)

def export_to_table(admin_id):
    conn = get_db()
    c = conn.cursor(cursor_factory=RealDictCursor)
    c.execute("SELECT * FROM answers")
    rows = c.fetchall()
    conn.close()
    if not rows:
        send_message(admin_id, MESSAGES["no_data"])
        return
    out = io.StringIO()
    cols = list(rows[0].keys())
    writer = csv.DictWriter(out, fieldnames=cols)
    writer.writeheader()
    for r in rows:
        writer.writerow(dict(r))
    fname = "survey_export.csv"
    with open(fname, "w", encoding="utf-8-sig") as f:
        f.write(out.getvalue())
    out.close()
    try:
        upload = vk_api.VkUpload(vk_session)
        doc = upload.document_message(fname, title="Выгрузка анкет")
        att = f"doc{doc['owner_id']}_{doc['id']}"
        send_message(admin_id, "Вот выгрузка собранных анкет:", attachment=att)
    except Exception as e:
        send_message(admin_id, f"Ошибка при загрузке файла: {e}")
    finally:
        if os.path.exists(fname):
            os.remove(fname)

# ----------------- ОСНОВНАЯ ЛОГИКА -----------------

def ask_step(user_id, step_key, uni_page=0):
    if step_key == "institution":
        send_message(user_id, QUESTIONS[step_key])
    elif step_key in OPTIONS:
        send_message(user_id, QUESTIONS[step_key])
    else:
        send_message(user_id, QUESTIONS[step_key])

def handle_message(event):
    user_id = event.user_id
    text = event.text.strip()

    if text.lower() in ["/export", "/выгрузить"]:
        if user_id in ADMIN_IDS:
            export_to_table(user_id)
        else:
            send_message(user_id, MESSAGES["admin_only"])
        return

    step_index, uni_page, started = get_progress(user_id)

    if started == 0:
        if text == "Начать анкету":
            set_progress(user_id, 0, 0, 1)
            ask_step(user_id, STEPS[0])
        else:
            send_message(user_id, MESSAGES["welcome"])
        return

    step_key = STEPS[step_index]

    # 1. Выбор ВУЗа
    if step_key == "institution":
        if text == "← Назад":
            if uni_page > 0:
                set_progress(user_id, step_index, uni_page - 1, 1)
                ask_step(user_id, step_key, uni_page - 1)
            else:
                send_message(user_id, "Это первая страница. Выберите учебное заведение.")
            return
        elif text == "Далее →":
            max_page = (len(UNIVERSITIES) - 1) // ITEMS_PER_PAGE
            if uni_page < max_page:
                set_progress(user_id, step_index, uni_page + 1, 1)
                ask_step(user_id, step_key, uni_page + 1)
            else:
                send_message(user_id, "Это последняя страница. Выберите учебное заведение.")
            return
        elif text in UNIVERSITIES:
            save_answer(user_id, "institution", text)
            next_idx = step_index + 1
            set_progress(user_id, next_idx, 0, 1)
            ask_step(user_id, STEPS[next_idx])
            return
        else:
            send_message(user_id, "Пожалуйста, выберите учебное заведение из кнопок или используйте «← Назад» / «Далее →».")
            return

    # 2. Валидация контактов
    if step_key == "contacts":
        ok, value = validate_contact(text)
        if ok:
            save_answer(user_id, "contacts", value)
            next_idx = step_index + 1
            set_progress(user_id, next_idx, 0, 1)
            ask_step(user_id, STEPS[next_idx])
            return
        else:
            send_message(user_id, MESSAGES["invalid_contact"])
            return

    # 3. Множественный выбор
    if step_key in MULTI_STEPS:
        max_opt = {"employment_status": 6, "post_plans": 7, "help_needed": 8}[step_key]
        nums = parse_multi_choice(text, max_opt)
        if nums is None:
            send_message(user_id, MESSAGES["invalid_multi"].format(max_opt))
            return
        label = multi_to_text(step_key, nums)
        save_answer(user_id, STEP_TO_DB[step_key], label)
        next_idx = step_index + 1
        set_progress(user_id, next_idx, 0, 1)
        if next_idx >= len(STEPS):
            send_message(user_id, MESSAGES["finished"])
        else:
            ask_step(user_id, STEPS[next_idx])
        return

    # 4. Одиночный выбор (кнопки)
    if step_key in OPTIONS:
        opts_lower = [o.lower() for o in OPTIONS[step_key]]
        if text.lower() not in opts_lower:
            send_message(user_id, MESSAGES["invalid_choice"])
            return
        for o in OPTIONS[step_key]:
            if o.lower() == text.lower():
                save_answer(user_id, STEP_TO_DB[step_key], o)
                break
        next_idx = step_index + 1
        set_progress(user_id, next_idx, 0, 1)
        if next_idx >= len(STEPS):
            send_message(user_id, MESSAGES["finished"])
        else:
            ask_step(user_id, STEPS[next_idx])
        return

    # 5. Свободный ввод
    if len(text) < 2:
        send_message(user_id, "Пожалуйста, введите более развёрнутый ответ.")
        return
    save_answer(user_id, STEP_TO_DB[step_key], text)
    next_idx = step_index + 1
    set_progress(user_id, next_idx, 0, 1)
    if next_idx >= len(STEPS):
        send_message(user_id, MESSAGES["finished"])
    else:
        ask_step(user_id, STEPS[next_idx])

# ----------------- ЗАПУСК -----------------

def main():
    init_db()
    print("Бот запущен...")
    while True:
        try:
            for event in longpoll.listen():
                if event.type == VkEventType.MESSAGE_NEW and event.to_me:
                    handle_message(event)
        except Exception as e:
            print(f"Ошибка в цикле: {e}")
            import time
            time.sleep(3)

if __name__ == "__main__":
    main()
