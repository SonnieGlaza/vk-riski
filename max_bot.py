import os
import re
import json
import asyncio
import logging
import io
import csv
import psycopg2
from psycopg2.extras import RealDictCursor
import aiohttp
import ssl
import certifi  # <-- добавлено

logging.basicConfig(level=logging.DEBUG, format="%(asctime)s [MAX] %(levelname)s %(message)s")
log = logging.getLogger("max_bot")

# ================= НАСТРОЙКИ =================
MAX_TOKEN = os.getenv("MAX_TOKEN", "")
ADMIN_IDS = set(map(int, os.getenv("ADMIN_IDS", "").split(","))) if os.getenv("ADMIN_IDS") else set()
DATABASE_URL = os.getenv("DATABASE_URL", "")
BASE_URL = "https://platform-api2.max.ru"

if not MAX_TOKEN or not DATABASE_URL:
    raise ValueError("Не заданы MAX_TOKEN или DATABASE_URL")

PHONE_PATTERN = re.compile(r'^(\+7|7|8)?[\s\-]?\(?\d{3}\)?[\s\-]?\d{3}[\s\-]?\d{2}[\s\-]?\d{2}$')
EMAIL_PATTERN = re.compile(r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$')

def format_numbered_list(items, start_from=1, truncate=True):
    lines = []
    for i, item in enumerate(items, start=start_from):
        if truncate and len(item) > 80:
            item = item[:77] + "…"
        lines.append(f"{i} — {item}")
    return "\n".join(lines)

def to_db_id(chat_id):
    return -abs(int(chat_id))

# ----------------- БАЗА ДАННЫХ -----------------
def get_db():
    return psycopg2.connect(DATABASE_URL)

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS answers (
            user_id BIGINT PRIMARY KEY,
            fio TEXT, institution TEXT, specialty TEXT, study_group TEXT,
            course TEXT, form_of_study TEXT, contacts TEXT,
            employment_status TEXT, target_contract TEXT, experience TEXT,
            practice_eval TEXT, events TEXT, resume_status TEXT,
            interview_training TEXT, special_status TEXT, military TEXT,
            maternity TEXT, graduate TEXT, post_plans TEXT, help_needed TEXT
        )
    """)
    c.execute("ALTER TABLE answers ADD COLUMN IF NOT EXISTS consent_status BOOLEAN DEFAULT FALSE")
    c.execute("""
        CREATE TABLE IF NOT EXISTS progress (
            user_id BIGINT PRIMARY KEY,
            step_index INTEGER DEFAULT 0,
            uni_page INTEGER DEFAULT 0,
            started INTEGER DEFAULT 0
        )
    """)
    conn.commit()
    conn.close()

def get_progress(db_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT step_index, uni_page, started FROM progress WHERE user_id=%s", (db_id,))
    row = c.fetchone()
    conn.close()
    return (row[0], row[1], row[2]) if row else (0, 0, 0)

def set_progress(db_id, step_index, uni_page=0, started=1):
    conn = get_db()
    c = conn.cursor()
    c.execute(
        "INSERT INTO progress (user_id, step_index, uni_page, started) VALUES (%s,%s,%s,%s) "
        "ON CONFLICT (user_id) DO UPDATE SET step_index=%s, uni_page=%s, started=%s",
        (db_id, step_index, uni_page, started, step_index, uni_page, started)
    )
    conn.commit()
    conn.close()

def save_answer(db_id, field, value):
    cols = ["fio","institution","specialty","study_group","course","form_of_study",
            "contacts","employment_status","target_contract","experience",
            "practice_eval","events","resume_status","interview_training",
            "special_status","military","maternity","graduate","post_plans","help_needed"]
    if field not in cols:
        return
    conn = get_db()
    c = conn.cursor()
    c.execute("INSERT INTO answers (user_id) VALUES (%s) ON CONFLICT (user_id) DO NOTHING", (db_id,))
    c.execute(f"UPDATE answers SET {field}=%s WHERE user_id=%s", (value, db_id))
    conn.commit()
    conn.close()

# ----------------- ВУЗы -----------------
UNIVERSITIES = [
    "БПОУ УР «Воткинский промышленный техникум»",
    "БПОУ УР «Воткинский музыкально-педагогический колледж имени П.И. Чайковского»",
    "БПОУ УР «Воткинский машиностроительный техникум имени В.Г.Садовникова»",
    "АПОУ УР «Глазовский аграрно-промышленный техникум»",
    "БПОУ УР «Глазовский технический колледж»",
    "БПОУ «Дебесский политехникум»",
    "БПОУР «Игринский политехнический техникум»",
    "БПОУ УР «Ижевский торгово-экономический техникум»",
    "БПОУ УР «Ижевский монтажный техникум»",
    "БПОУ «Ижевский агростроительный техникум»",
    "ЧПОО «Нефтяной техникум»",
    "БПОУ УР «Ижевский политехнический колледж»",
    "БПОУ УР «Ижевский промышленно-экономический колледж»",
    "БПОУ УР «Ижевский машиностроительный техникум им. С.Н. Борина»",
    "БПОУ УР «Радиомеханический техникум имени В.А. Шутова»",
    "АПОУ УР «Экономико-технологический колледж»",
    "БПОУ УР «Асановский аграрно-технический техникум»",
    "АПОУ УР «Топливно-энергетический колледж»",
    "БПОУ УР «Ижевский техникум индустрии питания»",
    "КПОУ УР «Удмуртский республиканский колледж культуры»",
    "АНПОО «Международный Восточно-Европейский колледж»",
    "АПОУ УР «Техникум радиоэлектроники и информационных технологий им. А.В. Воскресенского»",
    "АПОУ УР «Республиканский медицинский колледж имени Героя Советского Союза Ф.А. Пушиной Министерства здравоохранения Удмуртской Республики»",
    "ПОЧУ «Ижевский техникум экономики, управления и права Удмуртпотребсоюза»",
    "АНПОО СПО «Ижевский финансово-юридический колледж»",
    "БПОУ УР «Удмуртский республиканский социально-педагогический колледж»",
    "АПОУ УР «Строительный техникум»",
    "ФГБОУ ВО «Ижевская государственная медицинская академия»",
    "КПОУ УР «Республиканский музыкальный колледж»",
    "БПОУ УР «Ижевский промышленно-экономический колледж» в г. Можга",
    "БПОУ УР «Можгинский педагогический колледж имени Т.К. Борисова»",
    "БПОУ УР «Можгинский агропромышленный колледж»",
    "БПОУ УР «Сарапульский политехнический техникум»",
    "БПОУ УР «Сарапульский многопрофильный колледж»",
    "БПОУ УР «Сарапульский колледж социально-педагогических технологий и сервиса»",
    "БПОУ УР «Сюмсинский техникум лесного и сельского хозяйства»",
    "БПОУ УР «Увинский профессиональный колледж»",
    "БПОУ УР «Ярский политехникум»",
    "ФГБОУ ВО «Приволжский государственный университет путей сообщения»",
    "БПОУ УР «Ижевский индустриальный техникум имени Евгения Фёдоровича Драгунова»",
    "БПОУ УР «Глазовский политехнический колледж»",
    "Министерство юстиции Российской Федерации",
    "ФГБОУ ВО «Удмуртский государственный университет»",
    "ФГБОУ ВО «Удмуртский государственный аграрный университет»",
    "ФГБОУ ВО «Ижевский государственный технический университет имени М.Т. Калашникова»",
    "БПОУ УР «Ижевский автотранспортный техникум»",
    "ФГБОУ ВО «Глазовский государственный инженерно-педагогический университет имени В. Г. Корененко»",
    "Сарапульский техникум машиностроения и информационных технологий"
]
ITEMS_PER_PAGE = 10

# ----------------- ШАГИ АНКЕТЫ -----------------
STEPS = [
    "fio", "consent", "institution", "specialty", "study_group",
    "course", "form_of_study", "contacts",
    "employment_status", "target_contract", "experience",
    "practice_eval", "events", "resume_status",
    "interview_training", "special_status", "military",
    "maternity", "graduate", "post_plans", "help_needed"
]
STEP_TO_DB = {s: s for s in STEPS}

QUESTIONS = {
    "fio": "Пожалуйста, укажите ваши фамилию, имя и отчество полностью:",
    "consent": (
        "Я, {fio}, на основании статей 9, 11 Федерального закона от 27 июля 2006 г. N 152-ФЗ "
        "\"О персональных данных\" в целях моей профессиональной ориентации даю свое согласие "
        "казенному учреждению Удмуртской Республики «Республиканский центр занятости населения» "
        "на автоматизированную, а также без использования средств автоматизации обработку своих "
        "персональных данных, включая сбор, систематизацию, накопление, хранение, уточнение "
        "(обновление, изменение), использование, обезличивание, блокирование, уничтожение "
        "персональных данных о моих фамилии, имени, отчестве, номере телефона, адресе электронной почты.\n\n"
        "Настоящее согласие действует в течение 1 года с даты анкетирования.\n\n"
        "Пожалуйста, подтвердите согласие, нажав кнопку ниже:"
    ),
    "institution": "Выберите ваше учебное заведение из списка (используйте «далее» / «назад» для пролистывания):",
    "specialty": "Укажите вашу специальность обучения:",
    "study_group": "Укажите номер вашей учебной группы:",
    "course": "Выберите ваш курс обучения:",
    "form_of_study": "Выберите форму обучения:",
    "contacts": "Укажите контактные данные — телефон или e-mail (например: +79991234567 или student@mail.ru):",
    "employment_status": "Ваш статус занятости прямо сейчас. Выберите один вариант, указав его номер:",
    "target_contract": "Есть ли у вас заключённый договор о целевом обучении с работодателем?",
    "experience": "Есть ли у вас опыт работы или оплачиваемой стажировки по основной или близкой к ней специальности?",
    "practice_eval": "Как вы в целом оцениваете результаты своих производственных практик у работодателей по специальности обучения?",
    "events": (
        "Участвовали ли вы в течение обучения в мероприятиях, которые помогают познакомиться с работодателями "
        "(ярмарки вакансий, дни карьеры, профтуры на предприятия, встречи с работодателями и т.д.)?"
    ),
    "resume_status": "Наличие резюме для поиска работы:",
    "interview_training": "Проходили ли вы занятия или тренинги по навыкам прохождения собеседования?",
    "special_status": "Есть ли у вас особый статус или жизненные обстоятельства?",
    "military": (
        "Планируется ли в отношении вас призыв на военную службу в ближайшее время "
        "(после окончания текущего года обучения)?"
    ),
    "maternity": (
        "Планируете ли вы уходить в отпуск по уходу за ребёнком "
        "в период обучения или сразу после окончания обучения (или продолжать уже начатый отпуск)?"
    ),
    "graduate": "Являетесь ли вы студентом выпускного курса (оканчиваете программу в текущем учебном году)?",
    "post_plans": "Ваши планы после выпуска. Выберите один или несколько вариантов, указав их номера через запятую (например: 1, 3):",
    "help_needed": "Какую помощь от Кадрового центра «Работа России» вы бы считали наиболее полезной? Выберите все подходящие варианты, указав их номера через запятую (например: 1, 2, 4):"
}

OPTIONS = {
    "consent": ["Да, я согласен(на)", "Нет, я не согласен(на)"],
    "course": ["1 курс", "2 курс", "3 курс", "4 курс", "5 курс"],
    "form_of_study": ["очная", "очно-заочная", "заочная"],
    "employment_status": [
        "Работаю по трудовому договору (в том числе по совместительству)",
        "Работаю по гражданско-правовому договору (договор подряда, услуг и т.п.)",
        "Являюсь самозанятым / ИП / учредителем юрлица",
        "Прохожу оплачиваемую стажировку / практику у работодателя",
        "Работаю временно (разовые подработки), не по специальности обучения",
        "Ничего из вышеперечисленного"
    ],
    "target_contract": ["да, договор о целевом обучении заключён", "нет, договора о целевом обучении нет"],
    "experience": [
        "да, есть опыт работы / оплачиваемой стажировки по основной или близкой специальности",
        "есть опыт работы только вне специальности обучения",
        "нет, опыта работы и оплачиваемых стажировок пока не было"
    ],
    "practice_eval": [
        "скорее доволен(льна) или полностью доволен(льна)",
        "скорее не доволен(льна) / совсем не доволен(льна) результатами практик"
    ],
    "events": [
        "да, за последний год участвовал(а) хотя бы в одном таком мероприятии",
        "участвовал(а), но более года назад",
        "нет, ещё ни разу не участвовал(а)"
    ],
    "resume_status": [
        "есть актуальное резюме, которым я пользуюсь или готов(а) пользоваться",
        "резюме есть, но оно устарело / резюме нет, я его не составлял(а)"
    ],
    "interview_training": ["да, проходил(а) одно или несколько таких мероприятий", "пока не проходил(а)"],
    "special_status": [
        "да, имею группу инвалидности",
        "отношусь к категории детей-сирот и детей, оставшихся без попечения родителей",
        "планирую переезд в другой регион / страну после окончания обучения",
        "ничего из вышеперечисленного"
    ],
    "military": ["да, планируется призыв", "нет / не подлежу призыву / вопрос уже решён (служба пройдена и др.)"],
    "maternity": ["да, планирую", "пока не планирую"],
    "graduate": ["да, я учусь на выпускном курсе", "нет, я не на выпускном курсе"],
    "post_plans": [
        "У меня есть подписанный трудовой договор (или договор на целевое обучение)",
        "Есть устная договорённость с работодателем, но без подписанных документов",
        "Прохожу стажировку",
        "Планирую организовать своё дело (самозанятость / ИП / учредитель юрлица)",
        "Планирую продолжить обучение (магистратура / аспирантура и пр.)",
        "Сейчас ищу работу",
        "Пока нет планов"
    ],
    "help_needed": [
        "Подбор актуальных вакансий с учётом специальности",
        "Тренинги по составлению резюме, подготовке к собеседованиям, сопроводительных писем",
        "Профтур-экскурсии на предприятия",
        "Подбор оплачиваемой стажировки",
        "Подбор работодателя для практики",
        "Заключение договора с работодателем на целевое обучение",
        "Помощь с ЕЦП «Работа России»",
        "Другое (укажите)"
    ]
}
MULTI_STEPS = ["post_plans", "help_needed"]

MESSAGES = {
    "welcome": (
        "👋 Здравствуйте!\n\n"
        "Я задам несколько коротких вопросов о вашем обучении и занятости. "
        "Это займёт примерно 5–7 минут. По итогу анкетирования кадровый центр "
        "«Работа России» поможет Вам в прохождении тестирования на определение "
        "склонностей к профессиям, в составлении грамотного резюме, "
        "а также в подборе подходящих вакансий.\n\n"
        "Нажмите кнопку «Начать анкету», чтобы приступить."
    ),
    "invalid_contact": (
        "Не удалось распознать контакт. Пожалуйста, введите:\n\n"
        "• Номер телефона в формате +79991234567 или 89991234567\n"
        "или\n"
        "• Адрес электронной почты в формате example@mail.ru"
    ),
    "invalid_number": "Пожалуйста, введите номер от 1 до {}.",
    "invalid_multi": "Пожалуйста, укажите номера вариантов через запятую (например: 1, 3, 5). Проверьте, что номера от 1 до {}.",
    "no_data": "Пока нет собранных анкет для выгрузки.",
    "admin_only": "Эта команда доступна только администраторам.",
    "finished": (
        "✅ Спасибо! Анкета заполнена.\n\n"
        "Предоставленная вами информация позволит нам детально проанализировать ситуацию "
        "и предложить оптимальное решение.\n\n"
        "Если хотите пройти анкету заново — нажмите кнопку ниже."
    ),
    "already_finished": (
        "✅ Вы уже заполнили анкету ранее. Если хотите пройти заново — нажмите кнопку ниже."
    )
}

EXPORT_HEADERS = {
    "user_id": "ID пользователя", "fio": "ФИО", "institution": "Учебное заведение",
    "specialty": "Специальность", "study_group": "Учебная группа", "course": "Курс",
    "form_of_study": "Форма обучения", "contacts": "Контакты",
    "employment_status": "Статус занятости", "target_contract": "Целевой договор",
    "experience": "Опыт работы", "practice_eval": "Оценка практик",
    "events": "Участие в мероприятиях", "resume_status": "Наличие резюме",
    "interview_training": "Тренинги по собеседованию", "special_status": "Особый статус",
    "military": "Призыв на военную службу", "maternity": "Отпуск по уходу за ребёнком",
    "graduate": "Выпускной курс", "post_plans": "Планы после выпуска", "help_needed": "Нужная помощь"
}

# ----------------- КЛАВИАТУРЫ -----------------
def make_keyboard(kb_type):
    if kb_type == "start":
        return [{"type": "inline_keyboard", "payload": {"buttons": [[{"type": "message", "text": "Начать анкету"}]]}}]
    if kb_type == "restart":
        return [{"type": "inline_keyboard", "payload": {"buttons": [[{"type": "message", "text": "🔄 Пройти заново"}]]}}]
    return None

# ----------------- HTTP КЛИЕНТ MAX API -----------------
async def api_send_message(session, chat_id, text, attachments=None, user_id=None):
    url = f"{BASE_URL}/messages"
    params = {}
    if user_id:
        params["user_id"] = user_id
    else:
        params["chat_id"] = chat_id
    headers = {"Authorization": MAX_TOKEN, "Content-Type": "application/json"}

    MAX_TEXT = 4000
    chunks = [text]
    if len(text) > MAX_TEXT:
        chunks = []
        remaining = text
        while remaining:
            if len(remaining) <= MAX_TEXT:
                chunks.append(remaining)
                break
            split_pos = remaining.rfind("\n", 0, MAX_TEXT)
            if split_pos == -1:
                split_pos = MAX_TEXT
            chunks.append(remaining[:split_pos])
            remaining = remaining[split_pos:].lstrip("\n")

    for i, chunk in enumerate(chunks):
        body = {"text": chunk}
        if attachments and i == len(chunks) - 1:
            body["attachments"] = attachments
        try:
            async with session.post(url, params=params, headers=headers, json=body) as resp:
                resp_text = await resp.text()
                if resp.status != 200:
                    log.error("POST /messages -> %s: %s", resp.status, resp_text)
                    if not user_id and "chat_id" in params:
                        log.info("Пробую отправить через user_id=%s", chat_id)
                        params2 = {"user_id": chat_id}
                        async with session.post(url, params=params2, headers=headers, json=body) as resp2:
                            resp2_text = await resp2.text()
                            if resp2.status != 200:
                                log.error("POST /messages (user_id) -> %s: %s", resp2.status, resp2_text)
                            else:
                                log.debug("Отправлено через user_id")
                else:
                    log.debug("Сообщение отправлено, status=%s", resp.status)
        except Exception as e:
            log.error("Ошибка отправки: %s", e)
        if len(chunks) > 1:
            await asyncio.sleep(0.5)

async def api_get_updates(session, marker=None):
    params = {"limit": 100, "timeout": 30}
    if marker is not None:
        params["marker"] = marker
    headers = {"Authorization": MAX_TOKEN}
    async with session.get(f"{BASE_URL}/updates", params=params, headers=headers, timeout=aiohttp.ClientTimeout(total=95)) as resp:
        resp_text = await resp.text()
        if resp.status != 200:
            log.error("GET /updates -> %s: %s", resp.status, resp_text)
            return {"updates": [], "marker": marker}
        return json.loads(resp_text)

async def api_delete_webhook(session):
    headers = {"Authorization": MAX_TOKEN}
    try:
        async with session.delete(f"{BASE_URL}/subscriptions", headers=headers) as resp:
            log.info("DELETE /subscriptions -> %s", resp.status)
    except Exception as e:
        log.warning("delete_webhook: %s", e)

async def api_upload_file(session, file_path, filename):
    headers = {"Authorization": MAX_TOKEN}
    with open(file_path, "rb") as f:
        data = aiohttp.FormData()
        data.add_field("file", f, filename=filename)
        async with session.post(f"{BASE_URL}/uploads?type=file", headers=headers, data=data) as resp:
            result = await resp.json()
            log.info("Upload result: %s", result)
            return result

# ----------------- ОТПРАВКА СООБЩЕНИЙ -----------------
async def send_message(session, chat_id, message, keyboard_type=None, user_id=None):
    attachments = make_keyboard(keyboard_type)
    await api_send_message(session, chat_id, message, attachments, user_id)

# ----------------- ВОПРОСЫ -----------------
async def ask_university_page(session, chat_id, db_id, page, user_id=None):
    start = page * ITEMS_PER_PAGE
    end = min(start + ITEMS_PER_PAGE, len(UNIVERSITIES))
    items = UNIVERSITIES[start:end]
    list_text = format_numbered_list(items, start_from=start + 1, truncate=True)
    nav = []
    if page > 0:
        nav.append("«назад» — предыдущая страница")
    if end < len(UNIVERSITIES):
        nav.append("«далее» — следующая страница")
    nav_text = "\n".join(nav) if nav else ""
    message = f"{QUESTIONS['institution']}\n\n{list_text}"
    if nav_text:
        message += f"\n\n{nav_text}"
    message += "\n\nВведите номер вашего учебного заведения."
    await send_message(session, chat_id, message, user_id=user_id)

async def ask_step(session, chat_id, db_id, step_key, uni_page=0, user_id=None):
    if step_key == "institution":
        await ask_university_page(session, chat_id, db_id, uni_page, user_id)
    elif step_key == "consent":
        conn = get_db()
        c = conn.cursor()
        c.execute("SELECT fio FROM answers WHERE user_id=%s", (db_id,))
        row = c.fetchone()
        conn.close()
        fio_text = row[0] if row and row[0] else "[ФИО не указано]"
        message = QUESTIONS["consent"].format(fio=fio_text)
        opts = OPTIONS["consent"]
        list_text = format_numbered_list(opts, truncate=False)
        hint = "Напишите номер выбранного варианта (1 или 2)."
        await send_message(session, chat_id, f"{message}\n\n{list_text}\n\n{hint}", user_id=user_id)
    elif step_key in OPTIONS:
        opts = OPTIONS[step_key]
        list_text = format_numbered_list(opts, truncate=False)
        if step_key in MULTI_STEPS:
            hint = "Напишите номера выбранных вариантов через запятую (например: 1, 3)."
        else:
            hint = "Напишите номер выбранного варианта (например: 1)."
        message = f"{QUESTIONS[step_key]}\n\n{list_text}\n\n{hint}"
        await send_message(session, chat_id, message, user_id=user_id)
    else:
        await send_message(session, chat_id, QUESTIONS[step_key], user_id=user_id)

async def advance_step(session, chat_id, db_id, step_index, user_id=None):
    next_idx = step_index + 1
    if next_idx >= len(STEPS):
        set_progress(db_id, next_idx, 0, 2)
        await send_message(session, chat_id, MESSAGES["finished"], keyboard_type="restart", user_id=user_id)
    else:
        set_progress(db_id, next_idx, 0, 1)
        await ask_step(session, chat_id, db_id, STEPS[next_idx], user_id=user_id)

# ----------------- ПАРСИНГ -----------------
def validate_contact(text):
    text = text.strip()
    if PHONE_PATTERN.match(text):
        digits = re.sub(r'\D', '', text)
        if len(digits) == 11 and digits.startswith('8'):
            digits = '7' + digits[1:]
        elif len(digits) == 10:
            digits = '7' + digits
        if len(digits) == 11:
            return True, '+7' + digits[-10:]
    if EMAIL_PATTERN.match(text):
        return True, text.lower()
    return False, None

def parse_single_number(text, max_val):
    text = text.strip()
    if text.isdigit():
        n = int(text)
        if 1 <= n <= max_val:
            return n
    return None

def parse_multi_numbers(text, max_val):
    try:
        parts = [p.strip() for p in text.split(",") if p.strip()]
        nums = [int(p) for p in parts]
        if any(n < 1 or n > max_val for n in nums):
            return None
        return nums
    except ValueError:
        return None

# ----------------- ВЫГРУЗКА -----------------
def generate_xlsx():
    conn = get_db()
    c = conn.cursor(cursor_factory=RealDictCursor)
    c.execute("SELECT * FROM answers")
    rows = c.fetchall()
    conn.close()
    if not rows:
        return None, None
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Анкеты"
    cols = list(rows[0].keys())
    header_font = Font(bold=True)
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=EXPORT_HEADERS.get(col_name, col_name))
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row_idx, r in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(cols, start=1):
            val = r[col_name]
            ws.cell(row=row_idx, column=col_idx, value=val if val is not None else "")
    for col_idx, col_name in enumerate(cols, start=1):
        display_name = EXPORT_HEADERS.get(col_name, col_name)
        max_len = max(len(str(display_name)), max(
            (len(str(r[col_name])) if r[col_name] else 0) for r in rows
        ))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)
    fname = "survey_export_max.xlsx"
    wb.save(fname)
    return fname, rows

async def export_to_max(session, chat_id, user_id=None):
    fname, rows = generate_xlsx()
    if not rows:
        await send_message(session, chat_id, MESSAGES["no_data"], user_id=user_id)
        return
    try:
        result = await api_upload_file(session, fname, "survey_export.xlsx")
        if "token" not in result:
            raise Exception(f"Нет token в ответе: {result}")
        file_token = result["token"]
        attachments = [{"type": "file", "payload": {"token": file_token}}]
        await api_send_message(session, chat_id, "📊 Вот выгрузка анкет в Excel:", attachments, user_id)
    except Exception as e:
        log.error("Ошибка загрузки .xlsx: %s", e)
        try:
            out = io.StringIO()
            writer = csv.DictWriter(out, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            for r in rows:
                writer.writerow(dict(r))
            csv_text = out.getvalue()
            out.close()
            if len(csv_text) > 3900:
                chunks = []
                lines_csv = csv_text.split("\n")
                current = ""
                for line in lines_csv:
                    if len(current) + len(line) + 1 > 3900:
                        chunks.append(current)
                        current = line + "\n"
                    else:
                        current += line + "\n"
                if current:
                    chunks.append(current)
                for i, chunk in enumerate(chunks):
                    header = f"📊 Выгрузка анкет (часть {i+1}/{len(chunks)}):\n\n"
                    await send_message(session, chat_id, header + chunk, user_id=user_id)
            else:
                await send_message(session, chat_id, "📊 Выгрузка анкет (CSV):\n\n" + csv_text, user_id=user_id)
        except Exception as e2:
            await send_message(session, chat_id, f"Не удалось выгрузить данные: {e2}", user_id=user_id)
    finally:
        if os.path.exists(fname):
            os.remove(fname)

# ----------------- ОСНОВНАЯ ЛОГИКА -----------------
async def handle_message(session, chat_id, db_id, text, user_id=None):
    log.info("handle_message: chat_id=%s, db_id=%s, text=%r", chat_id, db_id, text[:50])

    if text.lower() in ["/export", "/выгрузить"]:
        if abs(db_id) in ADMIN_IDS:
            await export_to_max(session, chat_id, user_id)
        else:
            await send_message(session, chat_id, MESSAGES["admin_only"], user_id=user_id)
        return

    if text.lower() == "/restart":
        set_progress(db_id, 0, 0, 0)
        await send_message(session, chat_id, "Анкета сброшена. Нажмите «Начать анкету».", keyboard_type="start", user_id=user_id)
        return

    if text == "🔄 Пройти заново":
        set_progress(db_id, 0, 0, 0)
        await send_message(session, chat_id, MESSAGES["welcome"], keyboard_type="start", user_id=user_id)
        return

    step_index, uni_page, started = get_progress(db_id)

    if started == 0:
        if text == "Начать анкету":
            set_progress(db_id, 0, 0, 1)
            await ask_step(session, chat_id, db_id, STEPS[0], user_id=user_id)
        else:
            await send_message(session, chat_id, MESSAGES["welcome"], keyboard_type="start", user_id=user_id)
        return

    if started == 2 or step_index >= len(STEPS):
        await send_message(session, chat_id, MESSAGES["already_finished"], keyboard_type="restart", user_id=user_id)
        return

    step_key = STEPS[step_index]

    if step_key == "institution":
        if text.lower() in ["далее", ">", "следующий"]:
            max_page = (len(UNIVERSITIES) - 1) // ITEMS_PER_PAGE
            if uni_page < max_page:
                set_progress(db_id, step_index, uni_page + 1, 1)
                await ask_university_page(session, chat_id, db_id, uni_page + 1, user_id)
            else:
                await send_message(session, chat_id, "Это последняя страница.", user_id=user_id)
                await ask_university_page(session, chat_id, db_id, uni_page, user_id)
            return
        elif text.lower() in ["назад", "<", "←"]:
            if uni_page > 0:
                set_progress(db_id, step_index, uni_page - 1, 1)
                await ask_university_page(session, chat_id, db_id, uni_page - 1, user_id)
            else:
                await send_message(session, chat_id, "Это первая страница.", user_id=user_id)
                await ask_university_page(session, chat_id, db_id, uni_page, user_id)
            return
        if text.isdigit():
            idx = int(text) - 1
            if 0 <= idx < len(UNIVERSITIES):
                save_answer(db_id, "institution", UNIVERSITIES[idx])
                await advance_step(session, chat_id, db_id, step_index, user_id)
                return
        await send_message(session, chat_id, "Пожалуйста, введите номер учебного заведения из списка или используйте «далее» / «назад».", user_id=user_id)
        await ask_university_page(session, chat_id, db_id, uni_page, user_id)
        return

    if step_key == "contacts":
        ok, value = validate_contact(text)
        if ok:
            save_answer(db_id, "contacts", value)
            await advance_step(session, chat_id, db_id, step_index, user_id)
        else:
            await send_message(session, chat_id, MESSAGES["invalid_contact"], user_id=user_id)
        return

    if step_key in OPTIONS:
        opts = OPTIONS[step_key]
        if step_key == "consent":
            n = parse_single_number(text, len(opts))
            if n is None:
                await send_message(session, chat_id, MESSAGES["invalid_number"].format(len(opts)), user_id=user_id)
                return
            is_consent = (n == 1)
            conn = get_db()
            c = conn.cursor()
            c.execute("INSERT INTO answers (user_id) VALUES (%s) ON CONFLICT (user_id) DO NOTHING", (db_id,))
            c.execute("UPDATE answers SET consent_status=%s WHERE user_id=%s", (is_consent, db_id))
            conn.commit()
            conn.close()
            await advance_step(session, chat_id, db_id, step_index, user_id)
            return
        if step_key in MULTI_STEPS:
            nums = parse_multi_numbers(text, len(opts))
            if nums is None:
                await send_message(session, chat_id, MESSAGES["invalid_multi"].format(len(opts)), user_id=user_id)
                return
            label = "; ".join(opts[n - 1] for n in nums)
            save_answer(db_id, STEP_TO_DB[step_key], label)
        else:
            n = parse_single_number(text, len(opts))
            if n is None:
                await send_message(session, chat_id, MESSAGES["invalid_number"].format(len(opts)), user_id=user_id)
                return
            save_answer(db_id, STEP_TO_DB[step_key], opts[n - 1])
        await advance_step(session, chat_id, db_id, step_index, user_id)
        return

    if step_key not in ["post_plans", "help_needed"]:
        if len(text) < 2:
            await send_message(session, chat_id, "Пожалуйста, введите более развёрнутый ответ.", user_id=user_id)
            return

    save_answer(db_id, STEP_TO_DB[step_key], text)
    await advance_step(session, chat_id, db_id, step_index, user_id)

# ----------------- ОБРАБОТКА СОБЫТИЙ -----------------
async def handle_update(session, update):
    update_type = update.get("update_type", "")
    log.info("=== Получено событие: %s ===", update_type)
    log.debug("Полное событие: %s", json.dumps(update, ensure_ascii=False, indent=2))

    if update_type == "bot_started":
        chat_id = None
        user_id = None

        chat = update.get("chat", {})
        if chat:
            chat_id = chat.get("chat_id")
            chat_type = chat.get("type", "")
            log.info("bot_started: chat_id=%s, chat_type=%s", chat_id, chat_type)

        user = update.get("user", {})
        if user:
            user_id = user.get("user_id")
            log.info("bot_started: user_id=%s", user_id)

        target = chat_id if chat_id else user_id
        if not target:
            log.error("bot_started: не найден ни chat_id, ни user_id!")
            return

        db_id = to_db_id(target)
        await send_message(session, target, MESSAGES["welcome"], keyboard_type="start", user_id=user_id if not chat_id else None)
        return

    if update_type == "message_created":
        message = update.get("message", {})
        body = message.get("body", {})
        text = body.get("text", "").strip()

        if not text:
            log.warning("message_created: пустой текст")
            return

        recipient = message.get("recipient", {})
        chat_id = recipient.get("chat_id")
        chat_type = recipient.get("chat_type", "")

        sender = message.get("sender", {})
        user_id = sender.get("user_id")

        log.info("message_created: text=%r, chat_id=%s, chat_type=%s, user_id=%s", text[:80], chat_id, chat_type, user_id)

        target = chat_id if chat_id else user_id
        if not target:
            log.error("message_created: не найден ни chat_id, ни user_id!")
            return

        send_user_id = user_id if not chat_id else None

        db_id = to_db_id(target)
        try:
            await handle_message(session, target, db_id, text, user_id=send_user_id)
        except Exception as e:
            log.error("Ошибка обработки сообщения: %s", e, exc_info=True)
        return

    log.info("Неизвестный тип события: %s", update_type)

# ----------------- ЗАПУСК -----------------
async def main():
    init_db()
    log.info("=== MAX бот запускается ===")
    log.info("Токен: %s...%s", MAX_TOKEN[:8], MAX_TOKEN[-4:])
    log.info("Base URL: %s", BASE_URL)

    # <-- СКАЧИВАЕМ СЕРТИФИКАТЫ МИНЦИФРЫ И ОБЪЕДИНЯЕМ С CERTIFI -->
    import urllib.request

    RU_ROOT_CA = "https://gu-st.ru/content/lending/russian_trusted_root_ca_pem.crt"
    RU_SUB_CA = "https://gu-st.ru/content/lending/russian_trusted_sub_ca_pem.crt"

    combined_certs = certifi.where()  # путь к стандартному cacert.pem

    try:
        log.info("Скачиваю сертификаты Минцифры...")
        root_ca_path = "/tmp/russian_trusted_root_ca.pem"
        sub_ca_path = "/tmp/russian_trusted_sub_ca.pem"

        urllib.request.urlretrieve(RU_ROOT_CA, root_ca_path)
        urllib.request.urlretrieve(RU_SUB_CA, sub_ca_path)

        # Объединяем certifi + Минцифры в один файл
        combined_path = "/tmp/combined_cacert.pem"
        with open(combined_path, "wb") as out:
            with open(certifi.where(), "rb") as f:
                out.write(f.read())
            with open(root_ca_path, "rb") as f:
                out.write(f.read())
            with open(sub_ca_path, "rb") as f:
                out.write(f.read())

        ssl_context = ssl.create_default_context(cafile=combined_path)
        log.info("Сертификаты Минцифры загружены и объединены с certifi")
    except Exception as e:
        log.warning("Не удалось скачать сертификаты Минцифры: %s. Использую только certifi.", e)
        ssl_context = ssl.create_default_context(cafile=certifi.where())

    connector = aiohttp.TCPConnector(ssl=ssl_context)

    async with aiohttp.ClientSession(connector=connector) as session:
        # Удаляем webhook с подробным логированием
        log.info("Удаляю webhook (DELETE /subscriptions)...")
        try:
            async with session.delete(
                f"{BASE_URL}/subscriptions",
                headers={"Authorization": MAX_TOKEN}
            ) as resp:
                log.info("DELETE /subscriptions -> статус %s", resp.status)
                resp_text = await resp.text()
                log.info("Ответ: %s", resp_text)
        except Exception as e:
            log.error("Ошибка при удалении webhook: %s", e)

        await asyncio.sleep(1)

        # Проверяем соединение запросом /me
        log.info("Проверяю соединение (GET /me)...")
        try:
            async with session.get(
                f"{BASE_URL}/me",
                headers={"Authorization": MAX_TOKEN}
            ) as resp:
                log.info("GET /me -> статус %s", resp.status)
                resp_text = await resp.text()
                log.info("Ответ /me: %s", resp_text[:500])
        except Exception as e:
            log.error("Ошибка GET /me: %s", e)

        log.info("=== Polling запущен. Ожидание сообщений... ===")

        marker = None
        poll_count = 0
        while True:
            try:
                poll_count += 1
                data = await api_get_updates(session, marker)

                new_marker = data.get("marker")
                if new_marker is not None:
                    marker = new_marker

                updates = data.get("updates", [])

                # Логируем каждый 10-й poll, даже если пусто
                if poll_count % 10 == 0:
                    log.info("Poll #%d: получено %d обновлений, marker=%s",
                             poll_count, len(updates), marker)

                if updates:
                    log.info("Получено %d обновлений", len(updates))

                for update in updates:
                    try:
                        await handle_update(session, update)
                    except Exception as e:
                        log.error("Ошибка обработки update: %s", e, exc_info=True)

            except asyncio.TimeoutError:
                continue
            except Exception as e:
                log.error("Ошибка polling: %s", e)
                await asyncio.sleep(5)

if __name__ == "__main__":
    asyncio.run(main())

